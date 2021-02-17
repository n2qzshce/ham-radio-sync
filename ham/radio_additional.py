import logging
import openpyxl

from ham import radio_types
from ham.dmr.dmr_contact import DmrContact
from ham.dmr.dmr_id import DmrId
from ham.dmr.dmr_user import DmrUser
from ham.radio_channel import RadioChannel
from ham.radio_zone import RadioZone


class RadioAdditionalData:
	def __init__(self, channels, dmr_ids, digital_contacts, zones, users):
		self._channels = channels
		self._dmr_ids = dmr_ids
		self._digital_contacts = digital_contacts
		self._zones = zones
		self._users = users
		return

	def output(self, style):
		switch = {
			radio_types.DEFAULT: self._output_default,
			radio_types.BAOFENG: lambda x: False,
			radio_types.FTM400: lambda x: False,
			radio_types.D878: self._output_d878,
			radio_types.CS800: self._output_cs800
		}

		return switch[style](style)

	def _output_default(self, style):
		self._output_radioids_default(style)
		self._output_contacts_default(style)
		self._output_zones_default(style)
		self._output_user_default(style)

	def _output_radioids_default(self, style):
		if self._dmr_ids is None:
			logging.error(f"No DMR ids found for {style}.")
			return
		radio_id_file = open(f'out/{style}/{style}_radioid.csv', 'w+')

		headers = DmrId.create_empty()
		radio_id_file.write(headers.headers(style) + '\n')
		for dmr_id in self._dmr_ids.values():
			radio_id_file.write(dmr_id.output(style) + '\n')

		radio_id_file.close()
		return

	def _output_contacts_default(self, style):
		if self._digital_contacts is None:
			logging.error(f"No digital contacts found for {style}.")
			return

		dmr_contact_file = open(f'out/{style}/{style}_contacts.csv', 'w+')

		headers = DmrContact.create_empty()
		dmr_contact_file.write(headers.headers(style) + '\n')
		for dmr_contact in self._digital_contacts.values():
			dmr_contact_file.write(dmr_contact.output(style) + '\n')

		dmr_contact_file.close()

	def _output_zones_default(self, style):
		if self._zones is None:
			logging.error(f"No zones list found for {style}.")
			return

		zone_file = open(f'out/{style}/{style}_zone.csv', 'w+')
		headers = RadioZone.create_empty()
		zone_file.write(headers.headers(style) + '\n')
		for zone in self._zones.values():
			zone_file.write(zone.output(style) + '\n')

		zone_file.close()

	def _output_user_default(self, style):
		if self._users is None:
			logging.error(f"No zones list found for {style}.")
			return

		users_file = open(f'out/{style}/{style}_user.csv', 'w+')
		headers = DmrUser.create_empty()
		users_file.write(headers.headers(style) + '\n')
		for zone in self._users.values():
			users_file.write(zone.output(style) + '\n')

		users_file.close()

	def _output_d878(self, style):
		self._output_radioids_d878(style)
		self._output_contacts_d878(style)
		self._output_zones_d878(style)
		self._output_user_d878(style)

	def _output_radioids_d878(self, style):
		if self._dmr_ids is None:
			logging.error(f"No DMR ids found for {style}.")
			return
		radio_id_file = open(f'out/{style}/{style}_radioid.csv', 'w+')

		headers = DmrId.create_empty()
		radio_id_file.write(headers.headers(style) + '\n')
		for dmr_id in self._dmr_ids.values():
			radio_id_file.write(dmr_id.output(style) + '\n')

		radio_id_file.close()
		return

	def _output_contacts_d878(self, style):
		if self._digital_contacts is None:
			logging.error(f"No digital contacts found for {style}.")
			return

		dmr_contact_file = open(f'out/{style}/{style}_talkgroup.csv', 'w+')

		headers = DmrContact.create_empty()
		dmr_contact_file.write(headers.headers(style) + '\n')
		for dmr_contact in self._digital_contacts.values():
			dmr_contact_file.write(dmr_contact.output(style) + '\n')

		dmr_contact_file.close()

	def _output_zones_d878(self, style):
		if self._zones is None:
			logging.error(f"No zones list found for {style}.")
			return

		zone_file = open(f'out/{style}/{style}_zone.csv', 'w+')
		headers = RadioZone.create_empty()
		zone_file.write(headers.headers(style) + '\n')
		for zone in self._zones.values():
			if not zone.has_channels():
				continue
			zone_file.write(zone.output(style) + '\n')

		zone_file.close()

	def _output_user_d878(self, style):
		if self._users is None:
			logging.error(f"No zones list found for {style}.")
			return

		users_file = open(f'out/{style}/{style}_user.csv', 'w+')
		headers = DmrUser.create_empty()
		users_file.write(headers.headers(style) + '\n')
		for zone in self._users.values():
			users_file.write(zone.output(style) + '\n')

		users_file.close()

	def _output_cs800(self, style):
		channels_workbook = openpyxl.Workbook()
		analog_sheet = channels_workbook.create_sheet('Analog Channel', 0)
		digital_sheet = channels_workbook.create_sheet('Digital Channel', 1)
		channels_workbook.remove_sheet(channels_workbook.get_sheet_by_name('Sheet'))

		header = RadioChannel.create_empty()
		header.is_digital = lambda: False
		analog_sheet.append(header.headers(radio_types.CS800))
		header.is_digital = lambda: True
		digital_sheet.append(header.headers(radio_types.CS800))

		analog_num = 1
		digital_num = 1
		for radio_channel in self._channels.values():
			if radio_channel.is_digital():
				digital_sheet.append(radio_channel.output(radio_types.CS800, digital_num))
				digital_num += 1
			else:
				analog_sheet.append(radio_channel.output(radio_types.CS800, analog_num))
				analog_num += 1
		channels_workbook.save(f'out/{style}/{style}_channels.xlsx')
		channels_workbook.close()
		return

