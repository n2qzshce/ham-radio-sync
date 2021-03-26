import csv
import logging

import openpyxl

from src.ham.radio.cs800.dmr_contact_cs800 import DmrContactCs800
from src.ham.radio.cs800.radio_channel_cs800 import RadioChannelCS800
from src.ham.radio.d878.dmr_contact_d878 import DmrContactD878
from src.ham.radio.d878.dmr_id_d878 import DmrIdD878
from src.ham.radio.d878.dmr_user_d878 import DmrUserD878
from src.ham.radio.default_radio.dmr_contact_default import DmrContactDefault
from src.ham.radio.default_radio.dmr_id_default import DmrIdDefault
from src.ham.radio.default_radio.dmr_user_default import DmrUserDefault
from src.ham.radio.radio_casted_builder import RadioChannelBuilder, DmrContactBuilder, DmrIdBuilder, DmrUserBuilder
from src.ham.radio.radio_zone import RadioZone
from src.ham.util import radio_types, file_util
from src.ham.util.file_util import RadioWriter, FileUtil


class RadioAdditionalData:
	def __init__(self, channels, dmr_ids, digital_contacts, zones, users):
		self._channels = channels
		self._dmr_ids = dmr_ids
		self._digital_contacts = digital_contacts
		self._zones = zones
		self._users = users
		return

	def output(self, style):
		switch = {
			radio_types.DEFAULT: self._output_default,
			radio_types.BAOFENG: lambda x: False,
			radio_types.FTM400: lambda x: False,
			radio_types.D878: self._output_d878,
			radio_types.CS800: self._output_cs800,
			radio_types.D710: self._output_d710,
		}

		return switch[style](style)

	def _output_default(self, style):
		self._output_radioids_default(style)
		self._output_contacts_default(style)
		self._output_zones_default(style)
		self._output_user_default(style)

	def _output_radioids_default(self, style):
		if self._dmr_ids is None:
			logging.error(f"No DMR ids found for {style}.")
			return

		writer = FileUtil.open_file(f'out/{style}/{style}_radioid.csv', 'w+')
		radio_id_file = csv.writer(writer, lineterminator='\n')

		headers = DmrIdDefault.create_empty()
		radio_id_file.writerow(headers.headers())
		for dmr_id in self._dmr_ids.values():
			casted_id = DmrIdBuilder.casted(dmr_id, style)
			radio_id_file.writerow(casted_id.output())

		writer.close()
		return

	def _output_contacts_default(self, style):
		if self._digital_contacts is None:
			logging.error(f"No digital contacts found for {style}.")
			return

		writer = FileUtil.open_file(f'out/{style}/{style}_contacts.csv', 'w+')
		dmr_contact_file = csv.writer(writer, lineterminator='\n')

		headers = DmrContactDefault.create_empty()
		dmr_contact_file.writerow(headers.headers())
		for dmr_contact in self._digital_contacts.values():
			casted_contact = DmrContactBuilder.casted(dmr_contact, style)
			dmr_contact_file.writerow(casted_contact.output())

		writer.close()

	def _output_zones_default(self, style):
		if self._zones is None:
			logging.error(f"No zones list found for {style}.")
			return

		writer = FileUtil.open_file(f'out/{style}/{style}_zone.csv', 'w+')
		zone_file = csv.writer(writer, lineterminator='\n')

		headers = RadioZone.create_empty()
		zone_file.writerow(headers.headers(style))
		for zone in self._zones.values():
			zone_file.writerow(zone.output(style))

		writer.close()

	def _output_user_default(self, style):
		if self._users is None:
			logging.error(f"No zones list found for {style}.")
			return

		writer = FileUtil.open_file(f'out/{style}/{style}_user.csv', 'w+')
		users_file = csv.writer(writer, lineterminator='\n')

		headers = DmrUserDefault.create_empty()
		users_file.writerow(headers.headers())
		for user in self._users.values():
			casted_user = DmrUserBuilder.casted(user.cols, user.number, style)
			users_file.writerow(casted_user.output())

		writer.close()

	def _output_d878(self, style):
		self._output_radioids_d878(style)
		self._output_contacts_d878(style)
		self._output_zones_d878(style)
		self._output_user_d878(style)

	def _output_radioids_d878(self, style):
		logging.info(f"Writing {style} radio IDs")
		if self._dmr_ids is None:
			logging.error(f"No DMR ids found for {style}.")
			return

		radio_id_file = RadioWriter(f'out/{style}/{style}_radioid.csv', '\r\n')

		headers = DmrIdD878.create_empty()
		radio_id_file.writerow(headers.headers())
		for dmr_id in self._dmr_ids.values():
			casted_id = DmrIdBuilder.casted(dmr_id, radio_types.D878)
			radio_id_file.writerow(casted_id.output())

		radio_id_file.close()
		return

	def _output_contacts_d878(self, style):
		logging.info(f"Writing {style} contacts")
		if self._digital_contacts is None:
			logging.error(f"No digital contacts found for {style}.")
			return

		dmr_contact_file = RadioWriter(f'out/{style}/{style}_talkgroup.csv', '\r\n')

		headers = DmrContactD878.create_empty()
		dmr_contact_file.writerow(headers.headers())
		for dmr_contact in self._digital_contacts.values():
			casted_contact = DmrContactBuilder.casted(dmr_contact, style)
			dmr_contact_file.writerow(casted_contact.output())

		dmr_contact_file.close()

	def _output_zones_d878(self, style):
		logging.info(f"Writing {style} zones")
		if self._zones is None:
			logging.error(f"No zones list found for {style}.")
			return

		zone_file = RadioWriter(f'out/{style}/{style}_zone.csv', '\r\n')

		headers = RadioZone.create_empty()
		zone_file.writerow(headers.headers(style))
		for zone in self._zones.values():
			if not zone.has_channels():
				continue
			zone_file.writerow(zone.output(style))

		zone_file.close()

	def _output_user_d878(self, style):
		logging.info(f"Writing {style} users")
		if self._users is None:
			logging.error(f"No zones list found for {style}.")
			return

		users_file = RadioWriter(f'out/{style}/{style}_digital_contacts.csv', '\n')

		headers = DmrUserD878.create_empty()
		users_file.writerow(headers.headers())

		rows_processed = 1
		for user in self._users.values():
			casted_user = DmrUserBuilder.casted(user.cols, user.number, style)
			users_file.writerow(casted_user.output())
			rows_processed += 1
			logging.debug(f"Writing user row {rows_processed}")
			if rows_processed % file_util.USER_LINE_LOG_INTERVAL == 0:
				logging.info(f"Writing user row {rows_processed}")

		users_file.close()

	def _output_cs800(self, style):
		self._output_cs800_channels(style)
		self._output_cs800_user(style)

	def _output_cs800_channels(self, style):
		logging.info(f"Writing {style} channels")
		channels_workbook = openpyxl.Workbook()
		analog_sheet = channels_workbook.create_sheet('Analog Channel', 0)
		digital_sheet = channels_workbook.create_sheet('Digital Channel', 1)
		channels_workbook.remove_sheet(channels_workbook.get_sheet_by_name('Sheet'))

		header = RadioChannelCS800.create_empty()
		header.is_digital = lambda: False
		analog_sheet.append(header.headers())
		header.is_digital = lambda: True
		digital_sheet.append(header.headers())

		analog_num = 1
		digital_num = 1
		for radio_channel in self._channels.values():
			casted_channel = RadioChannelBuilder.casted(radio_channel, radio_types.CS800)

			if casted_channel.is_digital():
				digital_sheet.append(casted_channel.output(digital_num))
				digital_num += 1
			else:
				analog_sheet.append(casted_channel.output(analog_num))
				analog_num += 1
		channels_workbook.save(f'out/{style}/{style}_channels.xlsx')
		channels_workbook.close()
		return

	def _output_cs800_user(self, style):
		logging.info(f"Writing {style} users")
		user_workbook = openpyxl.Workbook()
		dmr_contacts_sheet = user_workbook.create_sheet('DMR_Contacts', 0)
		user_workbook.remove_sheet(user_workbook.get_sheet_by_name('Sheet'))

		headers = DmrContactCs800.create_empty()
		dmr_contacts_sheet.append(headers.headers())

		number = 1
		for dmr_contact in self._digital_contacts.values():
			casted_contact = DmrContactBuilder.casted(dmr_contact, style)
			if casted_contact.name.fmt_val() == 'Analog':
				continue
			casted_contact.number._fmt_val = number
			dmr_contacts_sheet.append(casted_contact.output())
			number += 1

		logging.info(f"Writing DMR users for {style}")
		for dmr_user in self._users.values():
			dmr_user.number._fmt_val = number
			casted_user = DmrUserBuilder.casted(dmr_user.cols, number, style)
			dmr_contacts_sheet.append(casted_user.output())
			number += 1
			logging.debug(f"Writing user row {number}")
			if number % file_util.USER_LINE_LOG_INTERVAL == 0:
				logging.info(f"Writing user row {number}")

		logging.info("Saving workbook...")
		user_workbook.save(f'out/{style}/{style}_user.xlsx')
		logging.info("Save done.")
		user_workbook.close()
		return

	def _output_d710(self, style):
		return

