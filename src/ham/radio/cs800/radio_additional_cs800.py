import logging

import openpyxl

from src.ham.radio.cs800.dmr_contact_cs800 import DmrContactCs800
from src.ham.radio.cs800.radio_channel_cs800 import RadioChannelCS800
from src.ham.radio.radio_additional import RadioAdditional
from src.ham.radio.radio_casted_builder import RadioChannelBuilder, DmrContactBuilder, DmrUserBuilder
from src.ham.util import radio_types, file_util


class RadioAdditionalCs800(RadioAdditional):
	def __init__(self, channels, dmr_ids, digital_contacts, zones, users):
		super().__init__(channels, dmr_ids, digital_contacts, zones, users)

	def output(self):
		self._output_channels()
		self._output_user()

	def _output_channels(self):
		logging.info(f"Writing {radio_types.CS800} channels")
		channels_workbook = openpyxl.Workbook()
		analog_sheet = channels_workbook.create_sheet('Analog Channel', 0)
		digital_sheet = channels_workbook.create_sheet('Digital Channel', 1)
		channels_workbook.remove_sheet(channels_workbook.get_sheet_by_name('Sheet'))

		header = RadioChannelCS800.create_empty()
		header.is_digital = lambda: False
		analog_sheet.append(header.headers())
		header.is_digital = lambda: True
		digital_sheet.append(header.headers())

		analog_num = 1
		digital_num = 1
		for radio_channel in self._channels:
			casted_channel = RadioChannelBuilder.casted(radio_channel, radio_types.CS800)

			if casted_channel.is_digital():
				digital_sheet.append(casted_channel.output(None))
				digital_num += 1
			else:
				analog_sheet.append(casted_channel.output(None))
				analog_num += 1
		channels_workbook.save(f'out/{radio_types.CS800}/{radio_types.CS800}_channels.xlsx')
		channels_workbook.close()
		return

	def _output_user(self):
		logging.info(f"Writing {radio_types.CS800} users")
		user_workbook = openpyxl.Workbook()
		dmr_contacts_sheet = user_workbook.create_sheet('DMR_Contacts', 0)
		user_workbook.remove_sheet(user_workbook.get_sheet_by_name('Sheet'))

		headers = DmrContactCs800.create_empty()
		dmr_contacts_sheet.append(headers.headers())

		number = 1
		for dmr_contact in self._digital_contacts.values():
			casted_contact = DmrContactBuilder.casted(dmr_contact, radio_types.CS800)
			if casted_contact.name.fmt_val() == 'Analog':
				continue
			dmr_contacts_sheet.append(casted_contact.output(None))
			number += 1

		logging.info(f"Writing DMR users for {radio_types.CS800}")
		for dmr_user in self._users.values():
			casted_user = DmrUserBuilder.casted(dmr_user.cols, radio_types.CS800)
			dmr_contacts_sheet.append(casted_user.output(number))
			number += 1
			logging.debug(f"Writing user row {number}")
			if number % file_util.USER_LINE_LOG_INTERVAL == 0:
				logging.info(f"Writing user row {number}")

		logging.info(f"Saving {radio_types.CS800} workbook...")
		user_workbook.save(f'out/{radio_types.CS800}/{radio_types.CS800}_user.xlsx')
		logging.info("Save done.")
		user_workbook.close()
		return

